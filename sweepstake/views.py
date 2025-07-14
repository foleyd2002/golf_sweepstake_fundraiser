from django.shortcuts import render
from .models import Contestant, Pick
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import redirect
from .models import Player
from django.contrib import messages
import os
import re
from django.views.decorators.cache import cache_page

# Create your views here.

def assign_tied_ranks(ranked):
    ranked_with_ranks = []
    prev_score = None
    prev_odds = None
    rank = 0
    real_rank = 0
    for entry in ranked:
        real_rank += 1
        if entry['combined_score'] != prev_score or entry['combined_odds_val'] != prev_odds:
            rank = real_rank
        ranked_with_ranks.append({'rank': rank, 'entry': entry, 'combined_odds': entry['combined_odds']})
        prev_score = entry['combined_score']
        prev_odds = entry['combined_odds_val']
    return ranked_with_ranks

def leaderboard(request):
    from math import prod
    # Cache leaderboard for 2 minutes to handle high traffic
    @cache_page(120)
    def cached_leaderboard(request):
        contestants = []
        # Use prefetch_related for picks and select_related for player
        for contestant in Contestant.objects.prefetch_related('picks__player').all():
            picks = list(contestant.picks.all())
            if len(picks) == 3:
                combined_score = sum(pick.player.current_score for pick in picks)
            else:
                combined_score = None  # Incomplete picks
            odds_values = []
            for pick in picks:
                try:
                    num, denom = map(int, pick.player.odds.replace(' ', '').split('/'))
                    odds_values.append((num, denom))
                except Exception:
                    odds_values.append((1, 1))
            if len(odds_values) == 3:
                combined_num = prod([n for n, d in odds_values])
                combined_denom = prod([d for n, d in odds_values])
                if combined_denom != 0:
                    x = round(combined_num / combined_denom)
                    combined_odds = f"{x}/1"
                    combined_odds_val = combined_num / combined_denom
                else:
                    combined_odds = "-"
                    combined_odds_val = 0
            else:
                combined_odds = "-"
                combined_odds_val = 0
            contestants.append({
                'contestant': contestant,
                'picks': picks,
                'combined_score': combined_score,
                'combined_odds': combined_odds,
                'combined_odds_val': combined_odds_val,
            })
        ranked = sorted(
            [c for c in contestants if c['combined_score'] is not None],
            key=lambda x: (x['combined_score'], -x['combined_odds_val'])
        )
        ranked_with_ranks = assign_tied_ranks(ranked)
        unranked = [c for c in contestants if c['combined_score'] is None]
        context = {'ranked': ranked_with_ranks, 'unranked': unranked}
        return render(request, 'sweepstake/leaderboard.html', context)
    return cached_leaderboard(request)

@staff_member_required
def update_scores(request):
    players = Player.objects.all().order_by('name')
    if request.method == 'POST':
        for player in players:
            score_field = f'score_{player.id}'
            if score_field in request.POST:
                try:
                    player.current_score = int(request.POST[score_field])
                    player.save()
                except ValueError:
                    pass
        return redirect('leaderboard')
    return render(request, 'sweepstake/update_scores.html', {'players': players})

@staff_member_required
def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        from openpyxl import load_workbook
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        # If the active sheet is empty or missing headers, fallback to 'sorted' if it exists
        if ws.max_row < 2 and 'sorted' in wb.sheetnames:
            ws = wb['sorted']
        imported = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            # New format: Name, Payment Email, Created At, Phone Number, Select Golfer 1, Golfer 1 odds, Select Golfer 2, Golfer 2 odds, Select Golfer 3, Golfer 3 odds
            name = row[0]
            payment_email = row[1]
            created_at = row[2]
            phone = row[3]
            golfer1_name = row[4]
            golfer1_odds = row[5]
            golfer2_name = row[6]
            golfer2_odds = row[7]
            golfer3_name = row[8]
            golfer3_odds = row[9]
            if not name:
                continue
            contestant, _ = Contestant.objects.get_or_create(
                name=name.strip(),
                created_at=str(created_at) if created_at else '',
                defaults={
                    'phone_number': str(phone) if phone else '',
                    'payment_email': payment_email or '',
                }
            )
            contestant.phone_number = str(phone) if phone else contestant.phone_number
            contestant.payment_email = payment_email or contestant.payment_email
            contestant.created_at = str(created_at) if created_at else contestant.created_at
            contestant.save()
            # Remove all existing picks for this contestant (entry) before adding new ones
            contestant.picks.all().delete()
            for golfer_name, odds in [
                (golfer1_name, golfer1_odds),
                (golfer2_name, golfer2_odds),
                (golfer3_name, golfer3_odds)
            ]:
                if not golfer_name:
                    continue
                odds_str = str(odds).strip() if odds else '100/1'
                player, _ = Player.objects.get_or_create(
                    name=str(golfer_name).strip(),
                    defaults={'odds': odds_str}
                )
                # Always update odds from Excel
                player.odds = odds_str
                player.save()
                Pick.objects.create(contestant=contestant, player=player)
            imported += 1
        messages.success(request, f"Imported/updated {imported} contestants.")
        return redirect('leaderboard')
    return render(request, 'sweepstake/import_excel.html')

@staff_member_required
def reset_leaderboard(request):
    if request.method == 'POST':
        # Delete all picks and contestants
        from .models import Pick, Contestant
        Pick.objects.all().delete()
        Contestant.objects.all().delete()
        # Optionally, also delete all Players:
        # from .models import Player
        # Player.objects.all().delete()
        messages.success(request, 'Leaderboard and all entries have been deleted.')
        return redirect('leaderboard')
    return render(request, 'sweepstake/reset_leaderboard.html')

def superuser_required(view_func):
    from django.contrib.auth.decorators import user_passes_test
    return user_passes_test(lambda u: u.is_active and u.is_superuser)(view_func)

@superuser_required
def reset_players(request):
    if request.method == 'POST':
        Player.objects.all().delete()
        messages.success(request, 'All golfers (players) have been deleted.')
        return redirect('admin_home')
    return render(request, 'sweepstake/reset_players.html')

def admin_home(request):
    # Only accessible to staff/superuser
    contestants = []
    for contestant in Contestant.objects.prefetch_related('picks__player').all():
        picks = list(contestant.picks.all())
        if len(picks) == 3:
            combined_score = sum(pick.player.current_score for pick in picks)
        else:
            combined_score = None
        contestants.append({
            'contestant': contestant,
            'entry_number': getattr(contestant, 'entry_number', None) or getattr(contestant, 'created_at', None),
            'picks': picks,
            'combined_score': combined_score,
        })
    ranked = [c for c in contestants if c['combined_score'] is not None]
    ranked.sort(key=lambda x: x['combined_score'])
    unranked = [c for c in contestants if c['combined_score'] is None]
    context = {
        'ranked': ranked,
        'unranked': unranked,
    }
    return render(request, 'sweepstake/admin_home.html', context)
